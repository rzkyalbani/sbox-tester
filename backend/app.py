from flask import Flask, jsonify, request, send_from_directory
import os
import json
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import os
import json
from typing import Dict, List

# Set up Flask app with proper static folder configuration
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
frontend_static = os.path.join(project_root, 'frontend', 'static')
app = Flask(__name__, static_folder=frontend_static, static_url_path='/static')

# Define directories for default and user S-boxes
DEFAULT_SBOXES_DIR = os.path.join(os.path.dirname(__file__), 'sboxes')
USER_SBOXES_DIR = os.path.join(os.path.dirname(__file__), 'user_sboxes')

# Create user sboxes directory if it doesn't exist
os.makedirs(USER_SBOXES_DIR, exist_ok=True)

@app.route('/')
def index():
    return send_from_directory('../frontend', 'dashboard.html')

@app.route('/sbox-analysis')
def sbox_analysis():
    return send_from_directory('../frontend', 'sbox-analysis.html')

@app.route('/text-encryption')
def text_encryption():
    return send_from_directory('../frontend', 'text-encryption.html')

@app.route('/image-encryption')
def image_encryption():
    return send_from_directory('../frontend', 'image-encryption.html')

@app.route('/api/sboxes', methods=['GET'])
def get_sboxes():
    """Returns list of available predefined S-boxes (both default and user)"""
    sboxes_info = []

    # Get default S-boxes
    default_sbox_files = [f for f in os.listdir(DEFAULT_SBOXES_DIR) if f.endswith('.json')]
    for filename in default_sbox_files:
        filepath = os.path.join(DEFAULT_SBOXES_DIR, filename)
        with open(filepath, 'r') as f:
            data = json.load(f)
            # Create an ID from the filename without extension
            sbox_id = os.path.splitext(filename)[0]
            sboxes_info.append({
                'id': sbox_id,
                'name': data.get('name', ''),
                'source': data.get('source', ''),
                'description': data.get('description', ''),
                'type': 'default'  # Indicate this is a default S-box
            })

    # Get user S-boxes
    user_sbox_files = [f for f in os.listdir(USER_SBOXES_DIR) if f.endswith('.json')]
    for filename in user_sbox_files:
        filepath = os.path.join(USER_SBOXES_DIR, filename)
        with open(filepath, 'r') as f:
            data = json.load(f)
            # Create an ID from the filename without extension
            sbox_id = os.path.splitext(filename)[0]
            sboxes_info.append({
                'id': sbox_id,
                'name': data.get('name', ''),
                'source': data.get('source', ''),
                'description': data.get('description', ''),
                'type': 'user'  # Indicate this is a user S-box
            })

    return jsonify(sboxes_info)

@app.route('/api/sbox/<sbox_id>', methods=['GET'])
def get_sbox_by_id(sbox_id):
    """Returns the content of a specific S-box by ID"""
    # First check in default S-boxes directory
    default_filepath = os.path.join(DEFAULT_SBOXES_DIR, f'{sbox_id}.json')
    if os.path.exists(default_filepath):
        with open(default_filepath, 'r') as f:
            data = json.load(f)
        return jsonify(data)

    # Then check in user S-boxes directory
    user_filepath = os.path.join(USER_SBOXES_DIR, f'{sbox_id}.json')
    if os.path.exists(user_filepath):
        with open(user_filepath, 'r') as f:
            data = json.load(f)
        return jsonify(data)

    return jsonify({"error": f"S-box '{sbox_id}' not found"}), 404

@app.route('/api/compute', methods=['POST'])
def compute_metrics():
    """Computes all S-box cryptographic metrics"""
    try:
        data = request.get_json()

        if not data or 'sbox' not in data:
            return jsonify({"ok": False, "error": "Missing 'sbox' in request body"}), 400

        sbox = data['sbox']

        # Validate input
        if not isinstance(sbox, list) or len(sbox) != 256:
            return jsonify({"ok": False, "error": "S-box must be a list of 256 integers"}), 400

        for val in sbox:
            if not isinstance(val, int) or val < 0 or val > 255:
                return jsonify({"ok": False, "error": "All values must be integers between 0 and 255"}), 400

        # Import metric functions
        from sbox_metrics.nl import compute_nl
        from sbox_metrics.sac import compute_sac
        from sbox_metrics.bic import compute_bic_nl, compute_bic_sac
        from sbox_metrics.lap import compute_lap
        from sbox_metrics.dap import compute_dap
        from sbox_metrics.du import compute_du
        from sbox_metrics.ad import compute_ad
        from sbox_metrics.to_metric import compute_to
        from sbox_metrics.ci import compute_ci

        # Calculate metrics
        metrics = {
            "NL": compute_nl(sbox),
            "SAC": compute_sac(sbox),
            "BIC_NL": compute_bic_nl(sbox),
            "BIC_SAC": compute_bic_sac(sbox),
            "LAP": compute_lap(sbox),
            "DAP": compute_dap(sbox),
            "DU": compute_du(sbox),
            "AD": compute_ad(sbox),
            "TO": compute_to(sbox),
            "CI": compute_ci(sbox)
        }

        return jsonify({"ok": True, "metrics": metrics})

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

# Import affine generator
from affine_generator import generate_affine_sbox, get_predefined_matrices

@app.route('/api/generate-affine-sbox', methods=['POST'])
def generate_affine():
    """Generate S-box using affine transformation"""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({"ok": False, "error": "Missing request body"}), 400
        
        # Get matrix value (can be hex string or integer)
        matrix = data.get('matrix', '0x57')
        if isinstance(matrix, str):
            matrix_value = int(matrix, 16) if matrix.startswith('0x') else int(matrix)
        else:
            matrix_value = int(matrix)
        
        # Get additive constant (default: 0x63 for standard AES)
        constant = int(data.get('constant', 0x63))
        
        # Validate inputs
        if not (0 <= matrix_value <= 255):
            return jsonify({"ok": False, "error": "Matrix value must be between 0 and 255"}), 400
        
        if not (0 <= constant <= 255):
            return jsonify({"ok": False, "error": "Additive constant must be between 0 and 255"}), 400
        
        # Generate S-box
        sbox = generate_affine_sbox(matrix_value, constant)
        
        return jsonify({
            "ok": True,
            "sbox": sbox,
            "matrix": hex(matrix_value),
            "constant": constant
        })
    
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route('/api/affine-matrices', methods=['GET'])
def get_affine_matrices():
    """Get list of predefined affine matrices"""
    try:
        matrices = get_predefined_matrices()
        return jsonify({"ok": True, "matrices": matrices})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route('/api/save-user-sbox', methods=['POST'])
def save_user_sbox():
    """Save a user-generated S-box to the user_sboxes directory"""
    try:
        data = request.get_json()

        if not data or 'sbox' not in data or 'name' not in data:
            return jsonify({"ok": False, "error": "Missing 'sbox' or 'name' in request body"}), 400

        sbox = data['sbox']
        name = data['name']
        description = data.get('description', '')
        source = data.get('source', 'User Generated')

        # Validate S-box
        if not isinstance(sbox, list) or len(sbox) != 256:
            return jsonify({"ok": False, "error": "S-box must be a list of 256 integers"}), 400

        for val in sbox:
            if not isinstance(val, int) or val < 0 or val > 255:
                return jsonify({"ok": False, "error": "All values must be integers between 0 and 255"}), 400

        # Validate name
        if not name or not isinstance(name, str):
            return jsonify({"ok": False, "error": "Name must be a non-empty string"}), 400

        # Create a safe filename from the name
        # Replace special characters and spaces with underscores
        import re
        safe_name = re.sub(r'[^a-zA-Z0-9_-]', '_', name)

        # Ensure the filename is not one of the default S-boxes
        if safe_name in [os.path.splitext(f)[0] for f in os.listdir(DEFAULT_SBOXES_DIR) if f.endswith('.json')]:
            return jsonify({"ok": False, "error": f"Cannot use name '{safe_name}' as it conflicts with a default S-box name"}), 400

        # Check if a user S-box with this name already exists
        filepath = os.path.join(USER_SBOXES_DIR, f'{safe_name}.json')
        if os.path.exists(filepath):
            return jsonify({"ok": False, "error": f"An S-box with the name '{name}' already exists"}), 400

        # Save the S-box to the user directory
        sbox_data = {
            "name": name,
            "source": source,
            "description": description,
            "sbox": sbox
        }

        with open(filepath, 'w') as f:
            json.dump(sbox_data, f, indent=2)

        return jsonify({
            "ok": True,
            "message": f"S-box '{name}' saved successfully",
            "sbox_id": safe_name
        })

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route('/api/delete-user-sbox/<sbox_id>', methods=['DELETE'])
def delete_user_sbox(sbox_id):
    """Delete a user-generated S-box from the user_sboxes directory"""
    try:
        # Validate sbox_id format
        import re
        if not re.match(r'^[a-zA-Z0-9_-]+$', sbox_id):
            return jsonify({"ok": False, "error": "Invalid S-box ID format"}), 400

        # Check if the S-box ID is a default S-box (should not be deletable)
        default_sbox_ids = [os.path.splitext(f)[0] for f in os.listdir(DEFAULT_SBOXES_DIR) if f.endswith('.json')]
        if sbox_id in default_sbox_ids:
            return jsonify({"ok": False, "error": f"Cannot delete default S-box '{sbox_id}'"}), 400

        # Check if the S-box exists in user directory
        filepath = os.path.join(USER_SBOXES_DIR, f'{sbox_id}.json')
        if not os.path.exists(filepath):
            return jsonify({"ok": False, "error": f"S-box '{sbox_id}' not found"}), 404

        # Delete the S-box file
        os.remove(filepath)

        return jsonify({
            "ok": True,
            "message": f"S-box '{sbox_id}' deleted successfully"
        })

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route('/api/sbox/upload-excel', methods=['POST'])
def upload_sbox_excel():
    """Upload S-box from Excel file"""
    try:
        if 'file' not in request.files:
            return jsonify({"ok": False, "error": "No file part in the request"}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({"ok": False, "error": "No file selected"}), 400

        if file and file.filename.lower().endswith(('.xlsx', '.xls')):
            # Read file content
            file_content = file.read()
            # Load workbook from content
            workbook = load_workbook(filename=io.BytesIO(file_content))
            worksheet = workbook.active

            # Read 256 values from the worksheet (16x16 grid)
            sbox = []
            for row in range(1, 17):  # 16 rows
                for col in range(1, 17):  # 16 columns
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value is None:
                        return jsonify({"ok": False, "error": f"Empty cell at row {row}, column {col}"}), 400
                    try:
                        val = int(cell_value)
                        if val < 0 or val > 255:
                            return jsonify({"ok": False, "error": f"Value {val} at row {row}, column {col} is out of range [0, 255]"}), 400
                        sbox.append(val)
                    except ValueError:
                        return jsonify({"ok": False, "error": f"Invalid value '{cell_value}' at row {row}, column {col}, must be an integer"}), 400

            if len(sbox) != 256:
                return jsonify({"ok": False, "error": "Excel file must contain exactly 256 values (16x16 grid)"}), 400

            # Validate that all values 0-255 are unique
            if len(set(sbox)) != 256:
                return jsonify({"ok": False, "error": "S-box must contain unique values"}), 400

            return jsonify({"ok": True, "sbox": sbox})

        else:
            return jsonify({"ok": False, "error": "Invalid file type. Please upload an Excel file (.xlsx or .xls)"}), 400

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route('/api/sbox/download-excel', methods=['POST'])
def download_sbox_excel():
    """Download S-box as Excel file"""
    try:
        data = request.get_json()

        if not data or 'sbox' not in data:
            return jsonify({"ok": False, "error": "Missing 'sbox' in request body"}), 400

        sbox = data['sbox']

        # Validate input
        if not isinstance(sbox, list) or len(sbox) != 256:
            return jsonify({"ok": False, "error": "S-box must be a list of 256 integers"}), 400

        for val in sbox:
            if not isinstance(val, int) or val < 0 or val > 255:
                return jsonify({"ok": False, "error": "All values must be integers between 0 and 255"}), 400

        # Create a new workbook and select the active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "S-Box Data"

        # Set up header
        ws['A1'] = "S-Box 16x16 Grid (Row = High Nibble, Column = Low Nibble)"
        ws['A1'].font = Font(bold=True)
        ws.merge_cells('A1:P1')  # Merge 16 columns for header

        # Add S-box values in a 16x16 grid starting from row 2, column 1
        for i, value in enumerate(sbox):
            row = (i // 16) + 2  # Start from row 2
            col = (i % 16) + 1   # Start from column 1
            ws.cell(row=row, column=col, value=value)

            # Optional: Add some styling
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Adjust column widths for better visualization
        for col in range(1, 17):  # Columns A to P
            ws.column_dimensions[get_column_letter(col)].width = 6

        # Create an in-memory buffer to hold the Excel file
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Return the Excel file as a response
        from flask import Response
        return Response(
            buffer.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=sbox.xlsx'}
        )

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

# Import crypto engine
from crypto_engine import SPNCipher, get_available_sboxes, get_sbox_by_id, validate_sbox

# Import image crypto engine
from image_crypto_engine import ImageSPNCipher, get_sbox_by_id as get_sbox_by_id_img, validate_sbox as validate_sbox_img

@app.route('/api/encrypt', methods=['POST'])
def encrypt_text():
    """Enkripsi teks menggunakan S-box tertentu"""
    try:
        data = request.get_json()

        if not data or 'plaintext' not in data or 'key' not in data or 'sbox_id' not in data:
            return jsonify({"ok": False, "error": "Missing 'plaintext', 'key', or 'sbox_id' in request body"}), 400

        plaintext = data['plaintext']
        key = data['key']
        sbox_id = data['sbox_id']

        # Ambil S-box berdasarkan ID
        sbox = get_sbox_by_id(sbox_id)

        # Validasi S-box
        if not validate_sbox(sbox):
            return jsonify({"ok": False, "error": "Invalid S-box format"}), 400

        # Buat cipher dan enkripsi
        cipher = SPNCipher(sbox)
        ciphertext = cipher.encrypt(plaintext, key)

        return jsonify({"ok": True, "ciphertext": ciphertext})

    except FileNotFoundError:
        return jsonify({"ok": False, "error": f"S-box '{sbox_id}' not found"}), 404
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route('/api/decrypt', methods=['POST'])
def decrypt_text():
    """Dekripsi teks menggunakan S-box tertentu"""
    try:
        data = request.get_json()

        if not data or 'ciphertext' not in data or 'key' not in data or 'sbox_id' not in data:
            return jsonify({"ok": False, "error": "Missing 'ciphertext', 'key', or 'sbox_id' in request body"}), 400

        ciphertext = data['ciphertext']
        key = data['key']
        sbox_id = data['sbox_id']

        # Ambil S-box berdasarkan ID
        sbox = get_sbox_by_id(sbox_id)

        # Validasi S-box
        if not validate_sbox(sbox):
            return jsonify({"ok": False, "error": "Invalid S-box format"}), 400

        # Buat cipher dan dekripsi
        cipher = SPNCipher(sbox)
        plaintext = cipher.decrypt(ciphertext, key)

        return jsonify({"ok": True, "plaintext": plaintext})

    except FileNotFoundError:
        return jsonify({"ok": False, "error": f"S-box '{sbox_id}' not found"}), 404
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route('/api/custom_encrypt', methods=['POST'])
def encrypt_with_custom_sbox():
    """Enkripsi teks menggunakan S-box kustom"""
    try:
        data = request.get_json()

        if not data or 'plaintext' not in data or 'key' not in data or 'sbox' not in data:
            return jsonify({"ok": False, "error": "Missing 'plaintext', 'key', or 'sbox' in request body"}), 400

        plaintext = data['plaintext']
        key = data['key']
        sbox = data['sbox']

        # Validasi S-box kustom
        if not validate_sbox(sbox):
            return jsonify({"ok": False, "error": "Invalid S-box format. Must contain 256 unique integers between 0-255."}), 400

        # Buat cipher dan enkripsi
        cipher = SPNCipher(sbox)
        ciphertext = cipher.encrypt(plaintext, key)

        return jsonify({"ok": True, "ciphertext": ciphertext})

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route('/api/custom_decrypt', methods=['POST'])
def decrypt_with_custom_sbox():
    """Dekripsi teks menggunakan S-box kustom"""
    try:
        data = request.get_json()

        if not data or 'ciphertext' not in data or 'key' not in data or 'sbox' not in data:
            return jsonify({"ok": False, "error": "Missing 'ciphertext', 'key', or 'sbox' in request body"}), 400

        ciphertext = data['ciphertext']
        key = data['key']
        sbox = data['sbox']

        # Validasi S-box kustom
        if not validate_sbox(sbox):
            return jsonify({"ok": False, "error": "Invalid S-box format. Must contain 256 unique integers between 0-255."}), 400

        # Buat cipher dan dekripsi
        cipher = SPNCipher(sbox)
        plaintext = cipher.decrypt(ciphertext, key)

        return jsonify({"ok": True, "plaintext": plaintext})

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# Endpoint untuk enkripsi/dekripsi gambar
@app.route('/api/image/encrypt', methods=['POST'])
def encrypt_image():
    """Enkripsi gambar menggunakan S-box tertentu"""
    try:
        if 'image' not in request.files:
            return jsonify({"ok": False, "error": "Missing 'image' file in request"}), 400

        if 'key' not in request.form or 'sbox_id' not in request.form:
            return jsonify({"ok": False, "error": "Missing 'key' or 'sbox_id' in request form"}), 400

        image_file = request.files['image']
        key = request.form['key']
        sbox_id = request.form['sbox_id']

        if image_file.filename == '':
            return jsonify({"ok": False, "error": "No image file selected"}), 400

        # Ambil S-box berdasarkan ID
        sbox = get_sbox_by_id_img(sbox_id)

        # Validasi S-box
        if not validate_sbox_img(sbox):
            return jsonify({"ok": False, "error": "Invalid S-box format"}), 400

        # Buat cipher dan enkripsi gambar
        cipher = ImageSPNCipher(sbox)

        # Baca file gambar ke buffer
        image_bytes = image_file.read()

        # Enkripsi gambar
        encrypted_image_bytes = cipher.encrypt_image_bytes_v2(image_bytes, key)

        # Kembalikan gambar terenkripsi
        from flask import Response
        return Response(
            encrypted_image_bytes,
            mimetype='image/png',
            headers={'Content-Disposition': 'attachment; filename=encrypted_image.png'}
        )

    except FileNotFoundError:
        return jsonify({"ok": False, "error": f"S-box '{sbox_id}' not found"}), 404
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route('/api/image/decrypt', methods=['POST'])
def decrypt_image():
    """Dekripsi gambar menggunakan S-box tertentu"""
    try:
        if 'image' not in request.files:
            return jsonify({"ok": False, "error": "Missing 'image' file in request"}), 400

        if 'key' not in request.form or 'sbox_id' not in request.form:
            return jsonify({"ok": False, "error": "Missing 'key' or 'sbox_id' in request form"}), 400

        image_file = request.files['image']
        key = request.form['key']
        sbox_id = request.form['sbox_id']

        if image_file.filename == '':
            return jsonify({"ok": False, "error": "No image file selected"}), 400

        # Ambil S-box berdasarkan ID
        sbox = get_sbox_by_id_img(sbox_id)

        # Validasi S-box
        if not validate_sbox_img(sbox):
            return jsonify({"ok": False, "error": "Invalid S-box format"}), 400

        # Buat cipher dan dekripsi gambar
        cipher = ImageSPNCipher(sbox)

        # Baca file gambar ke buffer
        image_bytes = image_file.read()

        # Dekripsi gambar
        decrypted_image_bytes = cipher.decrypt_image_bytes_v2(image_bytes, key)

        # Kembalikan gambar terdekripsi
        from flask import Response
        return Response(
            decrypted_image_bytes,
            mimetype='image/png',
            headers={'Content-Disposition': 'attachment; filename=decrypted_image.png'}
        )

    except FileNotFoundError:
        return jsonify({"ok": False, "error": f"S-box '{sbox_id}' not found"}), 404
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


if __name__ == '__main__':
    app.run(debug=True, host='127.0.0.1', port=5000)