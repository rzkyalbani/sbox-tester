import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Create a new workbook and select the active sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "S-Box Data"

# Create a simple S-box pattern (sequential values 0-255)
sbox = list(range(256))

# Add S-box values in a 16x16 grid starting from row 1, column 1 (no header)
for i, value in enumerate(sbox):
    row = (i // 16) + 1  # Start from row 1
    col = (i % 16) + 1   # Start from column 1
    ws.cell(row=row, column=col, value=value)

    # Optional: Add some styling
    cell = ws.cell(row=row, column=col)
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Adjust column widths for better visualization
for col in range(1, 17):  # Columns A to P
    ws.column_dimensions[get_column_letter(col)].width = 6

# Save the workbook
wb.save("test_sbox.xlsx")
print("Test Excel file 'test_sbox.xlsx' created successfully with a 16x16 S-box grid.")